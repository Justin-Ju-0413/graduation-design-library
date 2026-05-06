# -*- coding: utf-8 -*-
from openpyxl import load_workbook

input_file = "C:/Users/16084/Desktop/sample_courseEng_xq/2026042621592917.xlsx"
output_file = "C:/Users/16084/Desktop/sample_courseEng_xq/2026042621592917_filled.xlsx"

wb = load_workbook(input_file)
ws = wb.active

# 获取表头
headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col_idx).value
    if val:
        headers[val] = col_idx
        print(f"  列{col_idx}: {val}")

print(f"\n共 {ws.max_row - 1} 条数据")

# 课程映射: 中文课程名 -> (英文课程名, 英文分数, 英文学分, 英文学时, 英文学时单位, 英文课程类别)
course_mapping = {
    "线性代数与解析几何": ("Linear Algebra and Analytic Geometry", 77, 3.0, 48, "Hours", "RC"),
    "C++程序设计基础": ("C++ Programming Foundations", 77, 2.0, 40, "Hours", "RC"),
    "工程导论实践I": ("The Practice of Introduction to Engineering I", "A", 2.0, 2, "Weeks", "RC"),
    "微积分Ⅱ(一)": ("Calculus II", 62, 5.0, 80, "Hours", "RC"),
    "大学计算机基础": ("Foundations of Computer", 92, 0.0, 32, "Hours", "RC"),
    "学术英语与科技交流（一）": ("EAP & Technical Communication 1", 90, 3.0, 48, "Hours", "RC"),
    "体育(一)": ("Physical Education (1)", 94, 1.0, 36, "Hours", "RC"),
    "中国文学与文化": ("Chinese Classical Literature and Culture", 90, 2.0, 32, "Hours", "GE"),
    "工程导论I": ("Introduction to Engineering I", 85, 1.0, 16, "Hours", "RC"),
    "习近平新时代中国特色社会主义思想概论": (
        "An Introduction to Xi Jinping's Thoughts on Socialism with Chinese Characteristics in the New Era",
        75, 3.0, 48, "Hours", "RC"),
    "高级语言程序设计课程设计": ("Advanced Language Programming Course Design", "C", 1.0, 1, "Weeks", "RC"),
    "大学生心理健康教育": ("Psychological Health Education for College Student", 90, 2.0, 36, "Hours", "GE"),
    "电路实验": ("Experiment of Circuit", 88, 0.5, 16, "Hours", "RC"),
    "电子技术工程素质实践基础课": ("The Engineering Experiment of Electrical and Electronic", 92, 1.0, 1, "Weeks", "RC"),
    "马克思主义中国化进程与青年学生使命担当": (
        "Sinicization of Marxism and the Mission of Chinese Youth",
        78, 1.0, 20, "Hours", "GE"),
    "体育(二)": ("Physical Education (2)", 84, 1.0, 36, "Hours", "RC"),
    "微积分Ⅱ(二)": ("Calculus II", 64, 5.0, 80, "Hours", "RC"),
    "大学物理Ⅲ(一)": ("General Physics III(1)", 71, 4.0, 64, "Hours", "RC"),
    "建筑美学": ("Architectural Aesthetics", 87, 2.0, 32, "Hours", "GE"),
    "大学物理实验(一)": ("College Physical Experiment (I)", "A", 1.0, 32, "Hours", "RC"),
    "工程制图": ("Engineering Drawing", 81, 3.0, 48, "Hours", "RC"),
    "军事理论": ("Military Principle", 86, 2.0, 18, "Hours", "RC"),
    "思想道德与法治": ("Ethics and Rule of Law", 83, 2.5, 40, "Hours", "RC"),
    "电路": ("Electric Circuits", 60, 4.0, 64, "Hours", "RC"),
    "改革开放史": ("History of Reform and Opening-up", 90, 1.0, 16, "Hours", "GE"),
    "学术英语与科技交流（二）": ("EAP & Technical Communication 2", 92, 3.0, 48, "Hours", "RC"),
    "大学物理实验(二)": ("College Physical Experiment (II)", "C", 1.0, 32, "Hours", "RC"),
    "军事技能": ("Military Training", "A", 2.0, 2, "Weeks", "RC"),
    "马克思主义理论与实践": ("Marxism Theory and Practice", 83, 2.0, 2, "Weeks", "RC"),
    "中国近现代史纲要": ("An Outline of Chinese Near Past and Contemporary History", 79, 2.5, 40, "Hours", "RC"),
    "模拟电子技术实验": ("Experiment of Analog Electronics", 88, 0.5, 16, "Hours", "RC"),
    "体育(三)": ("Physical Education (3)", 90, 1.0, 32, "Hours", "RC"),
    "概率论与数理统计": ("Probability and Statistics", 60, 3.0, 48, "Hours", "RC"),
    "数字电子技术实验": ("Experiment of Digital Electronics", 90, 0.5, 16, "Hours", "RC"),
    "工程创新训练 II": ("Engineering Innovation Training II", 80, 2.0, 2, "Weeks", "RC"),
    "数字电子技术": ("Digital Electronics", 74, 4.0, 64, "Hours", "RC"),
    "微机系统与接口课程设计": ("Course Project of Microcomputer System and Interface Technology", 95, 1.0, 1, "Weeks", "RC"),
    "大学物理Ⅲ(二)": ("General Physics III(2)", 66, 4.0, 64, "Hours", "RC"),
    "毛泽东思想和中国特色社会主义理论体系概论": (
        "Introduction on Mao Zedong Thought and the Theoretical System of Socialism with Chinese Characteristics",
        76, 2.5, 40, "Hours", "RC"),
    "微机系统与接口": ("Microcomputer System and Interface Technology", 79, 3.5, 64, "Hours", "RC"),
    "信号与系统": ("Signals and Systems", 60, 4.0, 64, "Hours", "RC"),
    "信号与系统实验": ("Experiment of Signals and Systems", 91, 0.5, 16, "Hours", "RC"),
    "半导体物理": ("Semiconductor Physics", 61, 3.0, 48, "Hours", "RC"),
    "马克思主义基本原理": ("Introduction of the Marxism Basic Principle", 85, 2.5, 40, "Hours", "RC"),
    "模拟电子技术课程设计": ("Project of Analog Electronics", 88, 1.0, 1, "Weeks", "RC"),
    "体育(四)": ("Physical Education (4)", 87, 1.0, 36, "Hours", "RC"),
    "半导体物理与器件实验": ("Experiments on Semiconductor Physics and Devices", 88, 1.0, 32, "Hours", "RC"),
    "模拟电子技术": ("Analog Electronics", 60, 4.0, 64, "Hours", "RC"),
    "电子系统综合设计课程设计": ("Curriculum Design of the Synthetic Design of Electronic System", 81, 3.0, 80, "Hours", "EC"),
    "集成电路技术前沿": ("The Frontier of Integrated Circuit Technology", 90, 2.0, 32, "Hours", "EC"),
    "集成电路制造技术": ("IC Fabrication Technology", 60, 2.0, 32, "Hours", "RC"),
    "半导体器件": ("Semiconductor Devices", 61, 3.0, 48, "Hours", "RC"),
    "智能存算芯片系统架构及应用": ("Computational Memory/Storage System Architecture Design for AI Applications", 95, 2.0, 32, "Hours", "EC"),
    "微电子工艺实习": ("Practice of Microelectronics Process", 94, 2.0, 2, "Weeks", "RC"),
    "模拟集成电路原理与设计课程设计": ("Course Design of Analog Integrated Circuits", 95, 1.0, 1, "Weeks", "RC"),
    "Artificial Intelligence Foundations": ("Artificial Intelligence Foundations", 95, 2.0, 32, "Hours", "EC"),
    "模拟集成电路原理与设计": ("Analysis and Design of Analog Integrated Circuit", 60, 3.0, 48, "Hours", "RC"),
    "数字集成电路原理与设计课程设计": ("Course Design of Digital Integrated Circuits", 65, 1.0, 1, "Weeks", "RC"),
    "现代工程研究方法论": ("Methodology of Modern Engineering Research", 69, 2.0, 32, "Hours", "EC"),
    "电磁场与电磁波": ("Electromagnetic Fields and Waves", 68, 4.0, 64, "Hours", "EC"),
    "形势与政策": ("Analysis of the Situation & Policy", 67, 2.0, 64, "Hours", "RC"),
    "纳米器件与纳米电子学": ("Nano Devices and Nanoelectronics", 84, 2.0, 32, "Hours", "EC"),
    "微电子工艺创新实践": ("Innovation Research of Microelectronics Process", 80, 2.0, 2, "Weeks", "RC"),
    "毕业实习": ("Graduate Intern", 85, 4.0, 4, "Weeks", "RC"),
}

filled = 0
not_found = []

for row_idx in range(2, ws.max_row + 1):
    cn_name = ws.cell(row=row_idx, column=headers["中文课程名"]).value
    if cn_name and cn_name in course_mapping:
        eng_data = course_mapping[cn_name]
        ws.cell(row=row_idx, column=headers["英文课程名"], value=eng_data[0])
        ws.cell(row=row_idx, column=headers["英文分数"], value=eng_data[1])
        ws.cell(row=row_idx, column=headers["英文学分"], value=eng_data[2])
        ws.cell(row=row_idx, column=headers["英文学时"], value=eng_data[3])
        ws.cell(row=row_idx, column=headers["英文学时单位"], value=eng_data[4])
        ws.cell(row=row_idx, column=headers["英文课程类别"], value=eng_data[5])
        filled += 1
    elif cn_name:
        not_found.append(cn_name)

print(f"\n成功匹配: {filled} 条")
if not_found:
    print(f"未匹配: {not_found}")

wb.save(output_file)
print(f"\n保存完成: {output_file}")
